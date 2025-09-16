from io import BytesIO

from import_export.formats.base_formats import XLSX
from openpyxl import Workbook
from openpyxl.styles import Alignment


class CenterAlignedXLSX(XLSX):
    def export_data(self, dataset, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = dataset.title or "Export"

        alignment = Alignment(horizontal='center', vertical='center')

        column_widths = [len(header) for header in dataset.headers]

        ws.append(dataset.headers)
        for idx, cell in enumerate(ws[1], 0):
            cell.alignment = alignment

        for row_data in dataset.dict:
            row = [str(row_data.get(col, "")) for col in dataset.headers]
            ws.append(row)
            for col_idx, cell in enumerate(ws[ws.max_row], 0):
                cell.alignment = alignment
                value_length = len(cell.value) if cell.value else 0
                if value_length > column_widths[col_idx]:
                    column_widths[col_idx] = value_length

        for i, column_width in enumerate(column_widths):
            column_letter = ws.cell(row=1, column=i + 1).column_letter
            ws.column_dimensions[column_letter].width = column_width + 2

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()
